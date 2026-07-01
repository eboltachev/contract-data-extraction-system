from pathlib import Path
import pytest
from openpyxl import Workbook, load_workbook
from app.infrastructure.parsers.excel_template import read_criteria, fill_template
from app.infrastructure.repositories.file_repository import output_filename, FileRepository
from app.domain.extraction import ExtractionResult
from app.agents.base import BaseAgent
from app.domain.jobs import Job
from app.services.pipeline import Pipeline


def make_template(path: Path):
    wb=Workbook(); ws=wb.active; ws.title='Лист1'; ws.append(['№','Критерий','Значение']); ws.append([1,'Номер договора',None]); ws.append([2,'Дата подписания договора',None]); wb.save(path)

def test_read_criteria_and_columns(tmp_path):
    p=tmp_path/'t.xlsx'; make_template(p)
    _,_,val_col,criteria=read_criteria(p)
    assert val_col==3 and [c.name for c in criteria]==['Номер договора','Дата подписания договора']

def test_fill_template(tmp_path):
    p=tmp_path/'t.xlsx'; out=tmp_path/'out.xlsx'; make_template(p)
    fill_template(p,out,[ExtractionResult(criterion='Номер договора', value='ABC-1')])
    ws=load_workbook(out).active
    assert ws['C2'].value=='ABC-1'

def test_output_filename():
    assert output_filename('Договор №1.docx')=='Договор №1.xlsx'

@pytest.mark.asyncio
async def test_agent_iteration_fallback(monkeypatch):
    monkeypatch.setattr('app.core.config.settings.AGENT_MAX_ITERATIONS', 2, raising=False)
    a=BaseAgent(); a.max_iterations=2; calls=0
    async def nope():
        nonlocal calls; calls+=1; return None
    assert await a.bounded(nope, 'fallback')=='fallback' and calls==2

def test_fallback_value_present():
    r=ExtractionResult(criterion='X', value='Не найдено в договоре. Требует ручной проверки.')
    assert 'ручной проверки' in r.value

@pytest.mark.asyncio
async def test_rollback_on_stage_error(tmp_path):
    repo=type('R',(),{'job_dir':lambda self,j: tmp_path, 'save':lambda self,job: None})()
    async def save(job): pass
    repo.save=save
    files=FileRepository(); await files.prepare_dirs(tmp_path); (tmp_path/'output'/'partial.xlsx').write_text('x')
    job=Job(job_id='j')
    async def bad(): raise RuntimeError('boom')
    await Pipeline(repo, files).run(job, tmp_path/'missing.docx', tmp_path/'missing.xlsx')
    assert job.status=='failed'

@pytest.mark.asyncio
async def test_extraction_agent_deterministic_contract_fields():
    from app.agents.extraction import ExtractionAgent
    from app.domain.documents import DocumentFragment

    fragments = [
        DocumentFragment(section='header', text='Договор № СПД-01-15.04-25-ЛСПБ (25.2) ТМГ(1) от 15.04.2025'),
        DocumentFragment(section='preamble', text='ООО "Заказчик" и ООО "Ромашка", именуемое в дальнейшем "Исполнитель", заключили договор.'),
        DocumentFragment(section='Реквизиты сторон', text='ИНН 7701234567 КПП 770101001 ОГРН 1027700123456 БИК 044525225 р/с 40702810900000000001'),
        DocumentFragment(section='Приемка', text='Для закрытия договора Исполнитель передает акт КС-2, справку КС-3 и счет-фактуру.'),
        DocumentFragment(section='Сроки', text='Срок выполнения работ: в течение 30 календарных дней с даты подписания договора.'),
        DocumentFragment(section='Ответственность', text='За нарушение сроков начисляется пеня 0,1% от цены договора за каждый день просрочки.'),
    ]
    agent = ExtractionAgent()
    number = await agent.run_one('Номер договора', fragments)
    date = await agent.run_one('Дата подписания договора', fragments)
    counterparty = await agent.run_one('Контрагент', fragments)
    penalties = await agent.run_one('Штрафные санкции за нарушение договора', fragments)

    assert number.value == 'СПД-01-15.04-25-ЛСПБ (25.2) ТМГ(1)'
    assert date.value == '15.04.2025'
    assert 'Ромашка' in counterparty.value
    assert '0,1%' in penalties.value
