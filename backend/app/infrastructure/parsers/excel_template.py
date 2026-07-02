from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

from app.core.exceptions import ValidationError
from app.domain.criteria import Criterion
from app.domain.extraction import ExtractionResult


@dataclass(frozen=True)
class TemplateTable:
    worksheet: Worksheet
    header_row: int
    criterion_col: int
    value_col: int


CRITERION_HEADERS = {
    "критерий",
    "показатель",
    "параметр",
    "поле",
    "атрибут",
    "вопрос",
    "данные",
    "что извлечь",
    "наименование показателя",
    "требуемые сведения",
    "сведения",
}
VALUE_HEADERS = {
    "значение",
    "ответ",
    "результат",
    "извлеченное значение",
    "извлечённое значение",
    "заполнить",
    "данные из договора",
    "сведения из договора",
}
NUMBER_HEADERS = {"№", "no", "номер", "n"}


def find_template_table(workbook: Any):
    table = find_template_tables(workbook)[0]
    return table.worksheet, table.header_row, table.criterion_col, table.value_col


def find_template_tables(workbook: Any) -> list[TemplateTable]:
    tables: list[TemplateTable] = []
    for ws in workbook.worksheets:
        table = _find_explicit_table(ws)
        if table is not None:
            tables.append(table)

    if tables:
        return tables

    for ws in workbook.worksheets:
        table = _infer_headerless_table(ws)
        if table is not None:
            tables.append(table)

    if not tables:
        raise ValidationError(
            "В шаблоне не найдена таблица критериев. Поддерживаются заголовки "
            "типа: 'Критерий/Показатель/Параметр' и 'Значение/Ответ/Результат'."
        )

    return tables


def read_criteria(path: Path) -> tuple[object, object, int, list[Criterion]]:
    wb = load_workbook(path)
    tables = find_template_tables(wb)
    criteria: list[Criterion] = []

    for table in tables:
        for row_number in range(table.header_row + 1, table.worksheet.max_row + 1):
            raw = table.worksheet.cell(row_number, table.criterion_col).value
            name = _cell_text(raw)
            if not name:
                continue
            if _looks_like_service_row(name):
                continue
            criteria.append(
                Criterion(
                    row=row_number,
                    name=name,
                    sheet_name=table.worksheet.title,
                    criterion_col=table.criterion_col,
                    value_col=table.value_col,
                )
            )

    if not criteria:
        raise ValidationError("В шаблоне найдены колонки, но нет непустых критериев для извлечения.")

    first = tables[0]
    return wb, first.worksheet, first.value_col, criteria


def fill_template(template: Path, output: Path, results: list[ExtractionResult]) -> None:
    wb, ws, val_col, criteria = read_criteria(template)
    by_exact = {result.criterion: result for result in results}
    by_normalized = {_normalize(result.criterion): result for result in results}

    for criterion in criteria:
        target_ws = wb[criterion.sheet_name] if criterion.sheet_name else ws
        target_col = criterion.value_col or val_col
        cell = target_ws.cell(criterion.row, target_col)

        result = by_exact.get(criterion.name) or by_normalized.get(_normalize(criterion.name))
        cell.value = result.value if result else "Не найдено в договоре. Требует ручной проверки."
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        _fit_row_height(target_ws, criterion.row, str(cell.value))

    _write_audit_sheet(wb, results)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def _find_explicit_table(ws: Worksheet) -> TemplateTable | None:
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 30)):
        values = [_cell_text(cell.value) for cell in row]
        normalized = [_normalize(value) for value in values]

        criterion_idx = _first_header_index(normalized, CRITERION_HEADERS)
        value_idx = _first_header_index(normalized, VALUE_HEADERS)

        if criterion_idx is None:
            continue

        if value_idx is None:
            value_idx = _guess_value_col(normalized, criterion_idx)

        if value_idx is None or value_idx == criterion_idx:
            continue

        return TemplateTable(
            worksheet=ws,
            header_row=row[0].row,
            criterion_col=criterion_idx + 1,
            value_col=value_idx + 1,
        )

    return None


def _infer_headerless_table(ws: Worksheet) -> TemplateTable | None:
    best: tuple[int, int, int] | None = None

    for col in range(1, min(ws.max_column, 12) + 1):
        non_empty_text = 0
        empty_right = 0
        for row in range(1, min(ws.max_row, 80) + 1):
            left = _cell_text(ws.cell(row, col).value)
            right = _cell_text(ws.cell(row, col + 1).value) if col < ws.max_column else ""
            if left and len(left) > 4:
                non_empty_text += 1
                if not right:
                    empty_right += 1
        score = non_empty_text + empty_right
        if non_empty_text >= 2 and (best is None or score > best[0]):
            best = (score, col, col + 1)

    if best is None:
        return None

    return TemplateTable(worksheet=ws, header_row=0, criterion_col=best[1], value_col=best[2])


def _first_header_index(values: list[str], variants: set[str]) -> int | None:
    for idx, value in enumerate(values):
        if not value:
            continue
        if value in variants:
            return idx
        if any(value.startswith(variant + " ") or variant in value for variant in variants):
            return idx
    return None


def _guess_value_col(values: list[str], criterion_idx: int) -> int | None:
    for idx in range(criterion_idx + 1, len(values)):
        if values[idx] not in NUMBER_HEADERS:
            return idx
    return criterion_idx + 1 if criterion_idx + 1 < len(values) else None


def _fit_row_height(ws: Worksheet, row_number: int, value: str) -> None:
    lines = max(1, value.count("\n") + 1)
    approx_wrapped_lines = max(lines, len(value) // 85 + 1)
    current = ws.row_dimensions[row_number].height or 18
    ws.row_dimensions[row_number].height = max(current, min(220, 18 + approx_wrapped_lines * 13))


def _looks_like_service_row(value: str) -> bool:
    normalized = _normalize(value)
    return normalized in (CRITERION_HEADERS | VALUE_HEADERS | NUMBER_HEADERS)


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()


def _normalize(value: str) -> str:
    value = value.lower().replace("ё", "е")
    value = re.sub(r"[^а-яa-z0-9№%]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()



def _write_audit_sheet(wb: Any, results: list[ExtractionResult]) -> None:
    title = "Проверка"
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title=title)
    headers = ["Критерий", "Значение", "Нормализованное значение", "Источник", "Confidence", "Статус", "Комментарий"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for result in results:
        source = _source_text(result)
        status = _status(result)
        ws.append([
            result.criterion,
            result.value,
            result.normalized_value or "",
            source,
            round(float(result.confidence or 0.0), 3),
            status,
            result.reasoning_summary or "",
        ])

    widths = {"A": 36, "B": 64, "C": 28, "D": 72, "E": 12, "F": 18, "G": 56}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).alignment = Alignment(wrap_text=True, vertical="top")
        _fit_row_height(ws, row, str(ws.cell(row, 2).value or ""))


def _source_text(result: ExtractionResult) -> str:
    chunks: list[str] = []
    for fragment in result.source_fragments[:3]:
        clause = f", п. {fragment.clause}" if fragment.clause else ""
        text = _cell_text(fragment.text)
        chunks.append(f"{fragment.section or 'Документ'}{clause}: {text[:420]}")
    return "\n---\n".join(chunks)


def _status(result: ExtractionResult) -> str:
    if not result.value or result.value.startswith("Не найдено"):
        return "not_found"
    if result.value.startswith("Требует проверки") or result.confidence < 0.6:
        return "needs_review"
    return "verified"
